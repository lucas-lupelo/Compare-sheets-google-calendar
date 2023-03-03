import os
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
import pyautogui



def formatar_valor(valor):
    numero = valor.split(",")[0]
    decimal = valor.split(",")[1]
    if len(numero) == 4:
        numero_lista = list(numero)
        numero_lista.insert(1, ".")
        valor = "R$ " + str(''.join(numero_lista)) + "," + str(decimal)
    elif len(numero) < 4:
        valor = "R$ " + str(numero) + "," + str(decimal)
    elif len(numero) == 5:
        numero_lista = list(numero)
        numero_lista.insert(2, ".")
        valor = "R$ " + str(''.join(numero_lista)) + "," + str(decimal)
    elif len(numero) == 6:
        numero_lista = list(numero)
        numero_lista.insert(3, ".")
        valor = "R$ " + str(''.join(numero_lista)) + "," + str(decimal)
    elif len(numero) == 7:
        numero_lista = list(numero)
        numero_lista.insert(1, ".")
        numero_lista.insert(5, ".")
        valor = "R$ " + str(''.join(numero_lista)) + "," + str(decimal)
    elif len(numero) == 8:
        numero_lista = list(numero)
        numero_lista.insert(2, ".")
        numero_lista.insert(6, ".")
        valor = "R$ " + str(''.join(numero_lista)) + "," + str(decimal)
    elif len(numero) == 9:
        numero_lista = list(numero)
        numero_lista.insert(3, ".")
        numero_lista.insert(7, ".")
        valor = "R$ " + str(''.join(numero_lista)) + "," + str(decimal)
    return valor


#carrega todas as planilhas boleto da pasta
caminho = os.path.join(r"boletos_path")
arquivos_boletos = [os.path.join(caminho, nome_arquivo) for nome_arquivo in os.listdir(caminho) if nome_arquivo.endswith('.xlsx')]
planilha_boletos_multi = [load_workbook(wb) for wb in arquivos_boletos]
qtd_planilhas = len(planilha_boletos_multi)
# para acessar cada planilha planilha_boletos[0]["Boletos"].cell(row=2, column=3).value)



ativos_path = r"ativos_path"
planilha_ativos = load_workbook(ativos_path)
planilha_ativos = planilha_ativos['Ativos']


#17 coluna vencimento boletos
#3 coluna ativo boletos
#24 coluna banco boletos
#1 coluna cliente coleto

#2 coluna cliente ativos
#7 coluna saldo ativos
# 3 coluna ativo ativos
dados_individuais = []
dados_boletos = []
dados_ativos = []

# lista dados boletos
for planilha_num in range(0, qtd_planilhas):
    planilha_boletos = planilha_boletos_multi[planilha_num]["Boletos"]
    for linha in range(2, 1000):
        if planilha_boletos.cell(row=linha, column=3).value != None and planilha_boletos.cell(row=linha, column=5).value == "Aplicação" and planilha_boletos.cell(row=linha, column=17).value != "-":
            dados_individuais.append(planilha_boletos.cell(row=linha, column=3).value)
            dados_individuais.append(planilha_boletos.cell(row=linha, column=17).value)
            dados_individuais.append(planilha_boletos.cell(row=linha, column=24).value)
            dados_individuais.append(planilha_boletos.cell(row=linha, column=1).value) #cliente
            dados_boletos.append(dados_individuais)
            dados_individuais = []

#lista dados ativos
for linha in range(2, planilha_ativos.max_row):
    if planilha_ativos.cell(row=linha, column=7).value != '-':
        dados_individuais.append(planilha_ativos.cell(row=linha, column=2).value)
        dados_individuais.append(planilha_ativos.cell(row=linha, column=3).value)
        dados_individuais.append((str(planilha_ativos.cell(row=linha, column=7).value).split())[1].replace(".","").replace(",","."))
        dados_ativos.append(dados_individuais)
        dados_individuais = []


dados = []
#junção listas
for ativo_ativos in dados_ativos:
    for ativo_boletos in dados_boletos:
        if ativo_boletos[0] == ativo_ativos[1] and ativo_boletos[3] == ativo_ativos[0]:
            dados_individuais.append(ativo_ativos[0]) #cliente
            dados_individuais.append(ativo_ativos[1]) #ativo
            dados_individuais.append(ativo_boletos[2]) #banco
            dados_individuais.append(float(ativo_ativos[2])) #saldo
            if isinstance(ativo_boletos[1], datetime): #converter obtejo tipo datetime p/ string
                ativo_boletos[1] = ativo_boletos[1].strftime('%d/%m/%Y')
            dados_individuais.append(ativo_boletos[1]) #vencimento
            dados.append(dados_individuais)
            dados_individuais = []
            break

sublistas_unicas = set()
duplicatas = set()

for sublista in dados:
    # Converte a sublista em uma tupla, para que ela possa ser adicionada ao conjunto
    sublista_tupla = tuple(sublista)
    if sublista_tupla in sublistas_unicas:
        duplicatas.add(sublista_tupla)
    else:
        sublistas_unicas.add(sublista_tupla)

for duplicata in duplicatas:
    # Converte a tupla em uma lista novamente antes de remover a duplicata da lista original
    lista_duplicata = list(duplicata)
    dados.remove(lista_duplicata)


calendario_path = r"calendario_path"
calendario = load_workbook(calendario_path)
planilha_calendario = calendario['Vencimentos']
planilha_google = calendario['Google']

#gerar lista com o nome de todos os ativos, que servirá para não duplicar dados
duplicado = []
planilha_duplicados = pd.read_excel(calendario_path, sheet_name='Banco de dados')
num_linhas = len(planilha_duplicados.dropna()) #última linha preenchida

for linha in range(0, num_linhas):
    duplicado.append(planilha_duplicados.iloc[linha, 1])

planilha_calendario.delete_rows(2, planilha_calendario.max_row) #apaga dados da planilha para inserir apenas dados a serem atualizados
planilha_google.delete_rows(2, planilha_google.max_row)
#preenche planilha com os dados organizados na lista que unifica os dados das duas planilhas, ativos e boletos
planilha_duplicados = calendario["Banco de dados"] #carregar pelo openpyxl para preencher
linha = 2 #linha de escrita dos dados novos
num_linhas += 2 #linha de escrita no banco de dados (última)
for ativo in range(0, len(dados)):
    if dados[ativo][1] not in duplicado:
        for elemento in range(0, len(dados[ativo])):
            planilha_calendario.cell(row=linha, column=elemento + 1).value = dados[ativo][elemento]
            planilha_duplicados.cell(row=num_linhas, column=elemento + 1).value = dados[ativo][elemento]
        linha += 1
        num_linhas += 1

subject = "Vencimento ativo"
start_time = '10:00 AM'
all_day = "True"

cont = 2
for ativo in dados:
    if ativo[1] not in duplicado:
        descricao = f'Cliente: {ativo[0]}   |   Ativo: {ativo[1]}   |   Banco: {ativo[2]}   |   Saldo: {formatar_valor(str(ativo[3]).replace(".",","))}'
        print(descricao)
        start_date = ativo[4] #vencimento
        planilha_google.cell(row=cont, column=1).value = subject
        planilha_google.cell(row=cont, column=2).value = start_date
        planilha_google.cell(row=cont, column=3).value = start_time
        planilha_google.cell(row=cont, column=4).value = all_day
        planilha_google.cell(row=cont, column=5).value = descricao
        cont += 1

calendario.save(calendario_path)


read_file = pd.read_excel(calendario_path, sheet_name='Google')
read_file.to_csv(r"path_arquivo_final_gerado", index=False, header=True)

pyautogui.alert('Planilha atualizada.', button='OK', title='Atualizada')
